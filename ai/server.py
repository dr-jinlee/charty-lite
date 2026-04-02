"""
AI 서버
- 차트 생성 (Claude API)
- 파일 업로드 → STT (Whisper)
- 통역 (Claude API)
"""
import os
import re
import tempfile
import json
import time
import threading
import urllib.request
import numpy as np
import soundfile as sf
from flask import Flask, request, jsonify, send_from_directory, Response, stream_with_context
from flask_cors import CORS
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env'))

from chart_generator import generate_chart, generate_chart_stream, extract_summary
from interpreter import translate, translate_chart

app = Flask(__name__)
CORS(app)

# Whisper는 파일 업로드용으로만 사용 (지연 로딩)
whisper_model = None
_whisper_lock = threading.Lock()

def get_whisper():
    """Whisper 모델 지연 로딩 (스레드 안전)"""
    global whisper_model
    if whisper_model is None:
        with _whisper_lock:
            if whisper_model is None:
                from faster_whisper import WhisperModel
                model_size = os.getenv("WHISPER_MODEL", "medium")
                device = os.getenv("WHISPER_DEVICE", "cpu")
                compute_type = "float16" if device == "cuda" else "int8"
                print(f"[Whisper] 모델 로딩: {model_size} ({device})")
                whisper_model = WhisperModel(model_size, device=device, compute_type=compute_type)
                print(f"[Whisper] 로딩 완료")

    hints_path = os.path.join(os.path.dirname(__file__), "terminology", "dermatology_ko.json")
    hints = ""
    try:
        with open(hints_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            hints = data.get("whisper_hints", "")
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"[Whisper] 힌트 로드 실패: {e}")

    return whisper_model, hints


def _extract_json(raw: str) -> dict:
    """Claude 응답에서 JSON 안전 추출 (마크다운 코드블록 대응)"""
    raw = raw.strip()
    match = re.search(r'```(?:json)?\s*(.*?)\s*```', raw, re.DOTALL)
    if match:
        raw = match.group(1)
    return json.loads(raw)


print("=" * 50)
print("  Voice to Chart AI 서버")
print("  실시간 STT: 브라우저 Google STT")
print("  파일 STT: Whisper (지연 로딩)")
print("  차트 생성: Claude API")
print("=" * 50)


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "stt": "browser-google + whisper-file", "chart": "claude-api"})


_taxonomy_cache = {"data": None, "timestamp": 0}

def _fetch_taxonomy():
    """Procedure Hub taxonomy를 GitHub에서 가져오기 (5분 캐시)"""
    now = time.time()
    if _taxonomy_cache["data"] and (now - _taxonomy_cache["timestamp"]) < 300:
        return _taxonomy_cache["data"]
    try:
        url = "https://raw.githubusercontent.com/dr-jinlee/Procedure-Hub/main/src/data/taxonomy.json"
        req = urllib.request.Request(url, headers={"User-Agent": "Charty/1.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode())
        _taxonomy_cache["data"] = data
        _taxonomy_cache["timestamp"] = now
        print(f"[Taxonomy] 로드 완료: modality {len(data.get('modality', []))}개")
        return data
    except Exception as e:
        print(f"[Taxonomy] 로드 실패: {e}")
        return _taxonomy_cache["data"]  # 이전 캐시 반환


@app.route("/corrections", methods=["GET"])
def get_corrections():
    """STT 보정용 용어 매핑 (alias → 정식 용어) + Procedure Hub taxonomy"""
    terminology_path = os.path.join(os.path.dirname(__file__), "terminology", "dermatology_ko.json")
    try:
        with open(terminology_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"[보정] 용어 사전 로드 실패: {e}")
        data = {}

    corrections = {}

    # 1) 기존 용어 사전 (dermatology_ko.json)
    for section in ["procedures", "conditions"]:
        for items in data.get(section, {}).values():
            for item in items:
                for alias in item.get("aliases", []):
                    if alias and alias != item["term"]:
                        corrections[alias] = item["term"]

    for item in data.get("side_effects", []):
        for alias in item.get("aliases", []):
            if alias and alias != item["term"]:
                corrections[alias] = item["term"]

    for items in data.get("slang", {}).values():
        for item in items:
            if item.get("slang") and item.get("meaning"):
                corrections[item["slang"]] = item["meaning"]

    # 2) Procedure Hub taxonomy (제품 영문명 → 한글명, 서브카테고리 매칭)
    taxonomy = _fetch_taxonomy()
    if taxonomy:
        for cat in taxonomy.get("modality", []):
            # 제품 영문명 → 한글명 보정
            for prod in cat.get("products", []):
                name = prod.get("name", "")
                name_en = prod.get("name_en", "")
                # 영문명 → 한글명 (예: "Rejuran" → "리쥬란")
                if name_en and name and name_en.lower() != name.lower():
                    corrections[name_en] = name
                    # 소문자 버전도
                    corrections[name_en.lower()] = name
            # 서브카테고리 아이템의 detail → name 보정
            for sub in cat.get("subcategories", []):
                for item in sub.get("items", []):
                    detail = item.get("detail", "")
                    name = item.get("name", "")
                    # detail이 영문이면 한글 name으로 보정
                    if detail and name and detail != name and not any(ord(c) > 0x1100 for c in detail):
                        corrections[detail] = name

    # 3) 영어 발음 → 한글 음차 보정 (Google STT 한국어 모드에서 영어 발음 시)
    phonetic_corrections = {
        # 보톡스
        "보톡스": "보톡스", "보톡": "보톡스", "보택스": "보톡스", "보톡 스": "보톡스",
        "바톡스": "보톡스", "보탁스": "보톡스",
        # 필러
        "필러": "필러", "필라": "필러",
        # 쥬비덤
        "주비덤": "쥬비덤", "주비 덤": "쥬비덤", "쥬비 덤": "쥬비덤",
        "쥬비덤": "쥬비덤", "쥬비뎀": "쥬비덤", "주비뎀": "쥬비덤",
        # 레스틸렌
        "레스틸렌": "레스틸렌", "레스틸 렌": "레스틸렌", "레스틸란": "레스틸렌",
        "래스틸렌": "레스틸렌", "레스틸린": "레스틸렌",
        # 울쎄라
        "울세라": "울쎄라", "울떠라": "울쎄라", "얼떠라": "울쎄라",
        "울써라": "울쎄라", "울쎄라피": "울쎄라", "울떼라": "울쎄라",
        "얼세라": "울쎄라", "울테라": "울쎄라",
        # 써마지
        "서마지": "써마지", "떠마지": "써마지", "더마지": "써마지",
        "써머지": "써마지", "터마지": "써마지", "써마지에프엘엑스": "써마지FLX",
        # 리쥬란
        "레쥬란": "리쥬란", "리주란": "리쥬란", "레주란": "리쥬란",
        "리쮸란": "리쥬란", "레쮸란": "리쥬란", "리쥬 란": "리쥬란",
        # 쥬베룩
        "주베룩": "쥬베룩", "쥬베 룩": "쥬베룩", "주베 룩": "쥬베룩",
        "쥬벨룩": "쥬베룩", "주벨룩": "쥬베룩",
        # 피코
        "피코": "피코토닝", "피코 토닝": "피코토닝",
        # 엑셀V
        "엑셀 브이": "엑셀V", "엑셀브이": "엑셀V", "엑셀 비": "엑셀V",
        # 실펌
        "실펌 엑스": "실펌X", "실펌엑스": "실펌X",
        # 포텐자
        "포텐자": "포텐자", "포텐 자": "포텐자", "포텐져": "포텐자",
        # 스킨부스터
        "스킨 부스터": "스킨부스터", "스킨부스타": "스킨부스터",
        # 프로파일로
        "프로파일로": "프로파일로", "프로파일 로": "프로파일로", "프로파일러": "프로파일로",
        # 엑소좀
        "엑소 좀": "엑소좀", "엑소솜": "엑소좀", "엑소 솜": "엑소좀",
        # 인모드
        "인 모드": "인모드", "인모드": "인모드",
        # 슈링크
        "쉬링크": "슈링크", "슈링 크": "슈링크", "쉬링 크": "슈링크",
        # 엘란쎄
        "엘란세": "엘란쎄", "엘란 세": "엘란쎄", "엘란 쎄": "엘란쎄",
        # 스컬트라
        "스컬트라": "스컬트라", "스컬 트라": "스컬트라", "스컬프트라": "스컬트라",
        # 샤넬주사
        "샤넬 주사": "샤넬주사", "샤넬": "샤넬주사",
        # IPL
        "아이피엘": "IPL", "아이 피엘": "IPL", "아이피 엘": "IPL",
        # 제네시스
        "제네 시스": "제네시스", "제너시스": "제네시스",
        # 히알루론산
        "히알루론 산": "히알루론산", "히알루론산": "히알루론산",
        # PRP
        "피알피": "PRP", "피 알 피": "PRP",
    }
    for wrong, correct in phonetic_corrections.items():
        if wrong not in corrections and wrong != correct:
            corrections[wrong] = correct

    return jsonify({"corrections": corrections, "count": len(corrections)})


@app.route("/chart/generate", methods=["POST"])
def chart_generate():
    """상담 텍스트 → 차트 변환"""
    data = request.json
    transcript = data.get("transcript", "")
    consultant = data.get("consultant", "")
    consultation_type = data.get("consultationType", "auto")
    chart_style = data.get("chartStyle", "detailed")
    customer_profile = data.get("customerProfile")

    if not transcript.strip():
        return jsonify({"error": "상담 내용이 없습니다", "chart": "", "summary": ""})

    print(f"[Chart] 차트 생성 요청 - 텍스트 {len(transcript)}자")

    try:
        result = generate_chart(
            transcript=transcript,
            customer_profile=customer_profile,
            consultation_type=consultation_type,
            consultant=consultant,
            chart_style=chart_style,
        )
        print(f"[Chart] 생성 완료 - 차트 {len(result['chart'])}자")
        return jsonify(result)
    except Exception as e:
        print(f"[Chart] 생성 실패: {e}")
        return jsonify({"error": str(e), "chart": "", "summary": ""})


@app.route("/transcribe/file", methods=["POST"])
def transcribe_file():
    """오디오 파일 업로드 → Whisper로 텍스트 변환"""
    if 'audio' not in request.files:
        return jsonify({"error": "오디오 파일이 없습니다"})

    audio_file = request.files['audio']
    # 파일 크기 확인 (빈 파일 감지)
    audio_file.seek(0, 2)
    file_size = audio_file.tell()
    audio_file.seek(0)
    print(f"[Transcribe] 파일: {audio_file.filename}, 크기: {file_size} bytes, bilingual: {request.form.get('bilingual')}")
    if file_size < 1000:
        return jsonify({"error": f"오디오 파일이 너무 작습니다 ({file_size} bytes). 녹음이 제대로 되었는지 확인해주세요."})

    # 임시 파일로 저장
    suffix = os.path.splitext(audio_file.filename or '.wav')[1]
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        audio_file.save(tmp.name)
        tmp_path = tmp.name

    try:
        model, hints = get_whisper()
        is_bilingual = request.form.get("bilingual") == "true"

        print(f"[Transcribe] Whisper 변환 시작... (이중언어: {is_bilingual})")

        if is_bilingual:
            # 이중언어: language 미지정 → Whisper가 자동 감지
            segments, info = model.transcribe(
                tmp_path,
                language=None,
                vad_filter=True,
                vad_parameters=dict(min_silence_duration_ms=500, speech_pad_ms=300),
                beam_size=5,
                best_of=3,
            )
        else:
            segments, info = model.transcribe(
                tmp_path,
                language="ko",
                initial_prompt=hints,
                vad_filter=True,
                vad_parameters=dict(min_silence_duration_ms=600, speech_pad_ms=300),
                beam_size=5,
                best_of=3,
            )

        full_text = ""
        for seg in segments:
            full_text += seg.text

        full_text = full_text.strip()
        print(f"[Transcribe] 완료: {len(full_text)}자 (감지 언어: {info.language})")

        return jsonify({"transcript": full_text, "lang": info.language or "ko"})
    except Exception as e:
        print(f"[Transcribe] 실패: {e}")
        return jsonify({"error": str(e)})
    finally:
        os.unlink(tmp_path)


@app.route("/pricing", methods=["GET"])
def get_pricing():
    """가격 데이터 제공 (엑셀 업로드 > 크롤링 순으로 우선)"""
    base_dir = os.path.dirname(__file__)

    # 1순위: 엑셀 업로드 데이터
    manual_path = os.path.join(base_dir, "terminology", "clinic_pricing_manual.json")
    if os.path.exists(manual_path):
        try:
            with open(manual_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            data["source"] = "excel"
            return jsonify(data)
        except:
            pass

    # 2순위: 크롤링 데이터
    crawled_path = os.path.join(base_dir, "terminology", "clinic_pricing.json")
    try:
        with open(crawled_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        data["source"] = "crawling"
        return jsonify(data)
    except FileNotFoundError:
        return jsonify({"procedures": [], "source": "none"})


@app.route("/pricing/template", methods=["GET"])
def download_template():
    """엑셀 양식 다운로드"""
    static_dir = os.path.join(os.path.dirname(__file__), "static")
    return send_from_directory(static_dir, "가격표_양식.xlsx", as_attachment=True)


@app.route("/pricing/upload", methods=["POST"])
def upload_pricing():
    """엑셀 가격표 업로드 → JSON 변환 후 저장"""
    if "file" not in request.files:
        return jsonify({"error": "파일이 없습니다"}), 400

    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "엑셀 파일(.xlsx)만 지원합니다"}), 400

    try:
        from openpyxl import load_workbook
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        clinic_name = request.form.get("clinicName", "")

        # 헤더 행(2행)에서 열 위치 자동 감지 — 카테고리 열 삭제해도 동작
        headers = [str(cell.value or "").strip() for cell in ws[2]]
        cat_cols = [i for i, h in enumerate(headers) if h.startswith("카테고리")]
        name_col = next((i for i, h in enumerate(headers) if h == "시술명"), None)
        price_col = next((i for i, h in enumerate(headers) if "판매가" in h), None)
        orig_price_col = next((i for i, h in enumerate(headers) if "정가" in h), None)
        type_col = next((i for i, h in enumerate(headers) if "가격유형" in h), None)

        # 기존 양식 호환 (카테고리 1열 + 시술명 + 가격 구조)
        if name_col is None:
            name_col = 1 if not cat_cols else max(cat_cols) + 1
        if price_col is None:
            price_col = name_col + 1

        procedures = []

        for row in ws.iter_rows(min_row=3, values_only=True):
            # 카테고리 열들 합치기 (비어있는 건 건너뜀)
            cat_parts = []
            for ci in cat_cols:
                val = str(row[ci] or "").strip() if ci < len(row) else ""
                if val:
                    cat_parts.append(val)
            category = " > ".join(cat_parts) if cat_parts else str(row[0] or "").strip()

            name = str(row[name_col] or "").strip() if name_col < len(row) else ""
            price = row[price_col] if price_col < len(row) else None
            original_price = row[orig_price_col] if orig_price_col is not None and orig_price_col < len(row) else None
            price_type = str(row[type_col] or "정규가").strip() if type_col is not None and type_col < len(row) else "정규가"

            if not name or not price:
                continue
            # 숫자 변환 ("1,400,000원" → 1400000)
            def parse_price(v):
                if v is None:
                    return None
                if isinstance(v, (int, float)):
                    return int(v)
                s = str(v).replace(",", "").replace("원", "").replace(" ", "").strip()
                if not s:
                    return None
                try:
                    return int(float(s))
                except (ValueError, TypeError):
                    return None
            price = parse_price(price)
            original_price = parse_price(original_price)
            if not price:
                continue

            procedures.append({
                "category": category or "기타",
                "name": name,
                "price": price,
                "original_price": original_price,
                "price_type": price_type if price_type in ("정규가", "이벤트", "첫방문이벤트") else "정규가",
            })

        if not procedures:
            return jsonify({"error": "시술 데이터가 없습니다. 양식에 맞게 작성했는지 확인해주세요."}), 400

        # JSON으로 저장
        output = {
            "clinic_name": clinic_name or "엑셀 업로드",
            "uploaded_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
            "total_count": len(procedures),
            "procedures": procedures,
        }
        output_path = os.path.join(os.path.dirname(__file__), "terminology", "clinic_pricing_manual.json")
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)

        print(f"[가격표] 엑셀 업로드 완료: {len(procedures)}개 시술")
        return jsonify({"success": True, "count": len(procedures), "clinic_name": output["clinic_name"]})

    except Exception as e:
        print(f"[가격표] 업로드 실패: {e}")
        return jsonify({"error": f"파일 처리 실패: {str(e)}"}), 500


@app.route("/pricing/export", methods=["GET"])
def export_pricing():
    """현재 가격 데이터를 엑셀로 내보내기"""
    base_dir = os.path.dirname(__file__)

    # 엑셀 업로드 데이터 우선, 없으면 크롤링 데이터
    manual_path = os.path.join(base_dir, "terminology", "clinic_pricing_manual.json")
    crawled_path = os.path.join(base_dir, "terminology", "clinic_pricing.json")
    data = None
    for path in [manual_path, crawled_path]:
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                break
            except:
                pass

    if not data or not data.get("procedures"):
        return jsonify({"error": "내보낼 데이터가 없습니다"}), 404

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "시술 가격표"

        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # 헤더
        clinic_name = data.get("clinic_name", "")
        ws.merge_cells("A1:F1")
        ws["A1"] = f"{clinic_name} 시술 가격표"
        ws["A1"].font = Font(bold=True, size=12)
        ws.row_dimensions[1].height = 30

        headers = ["카테고리", "시술명", "판매가 (원)", "정가 (원)", "가격유형", "비고"]
        widths = [20, 50, 15, 15, 12, 20]
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[chr(64 + col)].width = w

        # 데이터
        for i, proc in enumerate(data["procedures"], 3):
            row = [
                proc.get("category", ""),
                proc.get("name", ""),
                proc.get("price", 0),
                proc.get("original_price"),
                proc.get("price_type", "정규가"),
                "",
            ]
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = thin_border
                if col in (3, 4) and val:
                    cell.number_format = '#,##0'

        # 임시 파일로 저장 후 전송
        export_path = os.path.join(tempfile.gettempdir(), "charty_export.xlsx")
        wb.save(export_path)
        print(f"[가격표] 내보내기: {len(data['procedures'])}개 시술")

        return send_from_directory(
            os.path.dirname(export_path),
            os.path.basename(export_path),
            as_attachment=True,
            download_name=f"가격표_{clinic_name or 'export'}.xlsx",
        )
    except Exception as e:
        print(f"[가격표] 내보내기 실패: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/pricing/reset", methods=["POST"])
def reset_manual_pricing():
    """엑셀 업로드 데이터 삭제 → 크롤링 데이터로 복원"""
    manual_path = os.path.join(os.path.dirname(__file__), "terminology", "clinic_pricing_manual.json")
    if os.path.exists(manual_path):
        os.remove(manual_path)
        print("[가격표] 엑셀 데이터 삭제 → 크롤링 데이터로 복원")
    return jsonify({"success": True})


# 환율 캐시 (1시간마다 갱신)
_exchange_cache = {"data": None, "timestamp": 0}

@app.route("/exchange-rates", methods=["GET"])
def get_exchange_rates():
    """실시간 환율 정보 (KRW 기준, 1시간 캐싱)"""
    now = time.time()
    # 캐시가 1시간 이내면 재사용
    if _exchange_cache["data"] and (now - _exchange_cache["timestamp"]) < 3600:
        return jsonify(_exchange_cache["data"])

    target_currencies = ["USD", "JPY", "CNY", "HKD", "TWD"]
    rates = {}

    try:
        # 무료 환율 API 사용
        url = "https://open.er-api.com/v6/latest/KRW"
        req = urllib.request.Request(url, headers={"User-Agent": "VoiceToChart/1.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            api_data = json.loads(resp.read().decode())

        if api_data.get("result") == "success":
            all_rates = api_data.get("rates", {})
            for currency in target_currencies:
                if currency in all_rates:
                    rates[currency] = all_rates[currency]

        print(f"[환율] 갱신 완료: {rates}")
    except Exception as e:
        print(f"[환율] API 실패: {e}, 폴백 사용")
        # API 실패 시 대략적인 폴백값
        rates = {
            "USD": 0.000714,   # ~1400원/달러
            "JPY": 0.107,      # ~9.3원/엔
            "CNY": 0.00518,    # ~193원/위안
            "HKD": 0.00557,    # ~180원/홍콩달러
            "TWD": 0.0232,     # ~43원/대만달러
        }

    result = {
        "base": "KRW",
        "rates": rates,
        "updated_at": time.strftime("%Y-%m-%d %H:%M", time.localtime()),
    }
    _exchange_cache["data"] = result
    _exchange_cache["timestamp"] = now
    return jsonify(result)


@app.route("/recommend", methods=["POST"])
def recommend_procedure():
    """실시간 시술 추천 (한 줄)"""
    data = request.get_json(silent=True)
    if not data or not data.get("transcript", "").strip():
        return jsonify({"recommendation": ""})

    transcript = data["transcript"][-600:]

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=40,
            system="너는 미용 클리닉 시술 추천 봇이다. 대화하지 마라. 시술명만 추천하라. 질문하지 마라. 설명하지 마라.",
            messages=[{"role": "user", "content": f"""아래 상담 녹취에서 언급된 시술과 시너지가 좋은 추가 시술 1개를 추천하라.

규칙:
- "OO도 같이 추천해보세요" 형식으로 딱 한 줄만 출력
- 25자 이내
- 이미 언급된 시술은 제외
- 추천할 게 없으면 빈 줄만 출력
- 절대 대화하지 마라

녹취: {transcript}

추천:"""}],
        )
        rec = response.content[0].text.strip().strip('"').strip("'")
        if len(rec) > 50:
            rec = ""
        print(f"[추천] {rec}")
        return jsonify({"recommendation": rec})
    except Exception as e:
        print(f"[추천] 실패: {e}")
        return jsonify({"recommendation": ""})


@app.route("/procedure-info", methods=["POST"])
def procedure_info():
    """시술 정보 AI 생성 (Procedure Hub에 없을 때 폴백)"""
    data = request.json
    keyword = data.get("keyword", "").strip()
    if not keyword:
        return jsonify({"error": "키워드 없음"})

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=500,
            messages=[{"role": "user", "content": f"""미용 클리닉 시술 "{keyword}"에 대해 아래 JSON 형식으로 간결하게 답해줘. 한국어로.

{{"overview": "한줄 설명", "mechanism": "작용 원리 (1-2문장)", "indications": {{"primary": ["적응증1", "적응증2", "적응증3"]}}, "safety": {{"common": ["흔한 부작용1", "흔한 부작용2"]}}, "onset_duration": {{"onset": "효과 발현 시간", "duration": "지속 기간"}}, "aftercare": ["시술 후 주의1", "시술 후 주의2"]}}

JSON만 출력해."""}],
        )
        result = _extract_json(response.content[0].text)
        print(f"[시술 AI] {keyword} 정보 생성 완료")
        return jsonify({"keyword": keyword, "extracted": result, "source": "ai"})
    except Exception as e:
        print(f"[시술 AI] 실패: {e}")
        return jsonify({"error": str(e)})


@app.route("/consultation/evaluate", methods=["POST"])
def evaluate_consultation():
    """상담 요약 평가 (AI 기반)"""
    data = request.json
    chart = data.get("chart", "")
    transcript = data.get("transcript", "")
    cart_items = data.get("cartItems", [])
    duration = data.get("duration", 0)

    if not chart:
        return jsonify({"error": "차트가 없습니다"})

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

        cart_text = "\n".join([f"- {item}" for item in cart_items]) if cart_items else "없음"
        minutes = max(1, round(duration / 60))

        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1200,
            messages=[{"role": "user", "content": f"""미용 클리닉 상담을 5가지 메트릭으로 평가해주세요.

## 상담 차트
{chart[:3000]}

## 상담 녹취 (일부)
{transcript[:1500]}

## 선택된 시술
{cart_text}

## 상담 시간
{minutes}분

## 5가지 평가 메트릭

1. **정확도** (0~100): 의학 지식/시술 정보가 정확한지. 잘못된 정보, 과장된 효과 설명이 없는지.
2. **친절도** (0~100): 어조, 뉘앙스, 톤 등 언어적 태도가 적절한지. 공감, 경청, 존중이 느껴지는지.
3. **세일즈** (0~100): 고객 고민에 맞는 적절한 시술을 추천했는지. 크로스셀링이 자연스러운지. 매출 기여도.
4. **신속도** (0~100): 횡설수설 없이 간결하고 효율적으로 상담했는지. 상담 시간 대비 내용 밀도.
5. **강요도** (0~100, 높을수록 좋음): 100=자연스러운 추천(이상적), 0=강압적 시술 강요(최악). 고객 의사 존중, 선택권 부여.

다음 JSON 형식으로만 답하세요:
{{"score": 82, "grade": "A", "metrics": [{{"name": "정확도", "score": 90, "emoji": "🎯", "comment": "시술 정보 정확함"}}, {{"name": "친절도", "score": 85, "emoji": "😊", "comment": "공감적 어조"}}, {{"name": "세일즈", "score": 75, "emoji": "💰", "comment": "추가 추천 부족"}}, {{"name": "신속도", "score": 88, "emoji": "⚡", "comment": "간결한 상담"}}, {{"name": "강요도", "score": 80, "emoji": "🤝", "comment": "자연스러운 추천"}}], "summary": "종합 평가 한줄", "strengths": ["잘한점1", "잘한점2"], "improvements": ["개선점1", "개선점2"]}}

score는 5개 메트릭 평균. grade: S(95+)/A(85+)/B(70+)/C(50+)/D(50미만). JSON만 출력."""}],
        )

        raw = response.content[0].text.strip()
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[1] if "\n" in raw else raw[3:]
        if raw.endswith("```"):
            raw = raw[:-3].strip()
        result = json.loads(raw)
        print(f"[평가] 상담 평가 완료: {result.get('grade', '?')}등급")
        return jsonify(result)
    except Exception as e:
        print(f"[평가] 실패: {e}")
        return jsonify({"error": str(e)})


@app.route("/consultation/report", methods=["POST"])
def generate_report():
    """상담 상세 평가 리포트 (Sonnet, 풍부한 내용)"""
    data = request.json
    chart = data.get("chart", "")
    transcript = data.get("transcript", "")
    cart_items = data.get("cartItems", [])
    duration = data.get("duration", 0)
    eval_summary = data.get("evalSummary", "")

    if not chart and not transcript:
        return jsonify({"error": "데이터 없음"})

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

        cart_text = "\n".join([f"- {item}" for item in cart_items]) if cart_items else "없음"
        minutes = max(1, round(duration / 60))

        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=2000,
            messages=[{"role": "user", "content": f"""미용 클리닉 상담에 대한 **상세 평가 리포트**를 작성해주세요.

## 상담 차트
{chart[:4000]}

## 상담 녹취
{transcript[:3000]}

## 선택 시술
{cart_text}

## 상담 시간
{minutes}분

## 간략 평가 결과
{eval_summary}

아래 JSON 형식으로 상세 리포트를 작성하세요:
{{
  "title": "상담 평가 리포트",
  "date": "2026-03-29",
  "duration_min": {minutes},
  "overall": {{
    "score": 85,
    "grade": "A",
    "one_liner": "한줄 총평"
  }},
  "metrics": [
    {{
      "name": "정확도",
      "score": 90,
      "emoji": "🎯",
      "analysis": "상세 분석 2~3문장. 어떤 정보가 정확했고, 어떤 부분이 부족했는지.",
      "examples": ["녹취에서 발견된 구체적 예시"],
      "suggestion": "구체적 개선 제안"
    }},
    {{
      "name": "친절도",
      "score": 85,
      "emoji": "😊",
      "analysis": "어조/뉘앙스 분석",
      "examples": ["좋았던 표현 or 아쉬운 표현"],
      "suggestion": "개선 제안"
    }},
    {{
      "name": "세일즈",
      "score": 75,
      "emoji": "💰",
      "analysis": "크로스셀링, 업셀링, 매출 기여 분석",
      "examples": ["추천 시술 예시"],
      "suggestion": "매출 향상 팁"
    }},
    {{
      "name": "신속도",
      "score": 88,
      "emoji": "⚡",
      "analysis": "상담 효율성 분석",
      "examples": ["간결했던/장황했던 부분"],
      "suggestion": "개선 제안"
    }},
    {{
      "name": "강요도",
      "score": 80,
      "emoji": "🤝",
      "analysis": "고객 의사 존중 수준 분석 (높을수록 좋음)",
      "examples": ["자연스러웠던/강압적이었던 부분"],
      "suggestion": "개선 제안"
    }}
  ],
  "highlights": ["특별히 잘한 점 1", "잘한 점 2", "잘한 점 3"],
  "risks": ["리스크/주의점 1", "리스크 2"],
  "action_items": ["다음 상담에서 할 것 1", "할 것 2", "할 것 3"],
  "coaching": "코칭 메시지 2~3문장. 격려와 구체적 성장 방향."
}}

JSON만 출력."""}],
        )

        raw = response.content[0].text.strip()
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[1] if "\n" in raw else raw[3:]
        if raw.endswith("```"):
            raw = raw[:-3].strip()
        result = json.loads(raw)
        print(f"[리포트] 상세 평가 리포트 생성 완료")
        return jsonify(result)
    except Exception as e:
        print(f"[리포트] 실패: {e}")
        return jsonify({"error": str(e)})


@app.route("/chart/generate-stream", methods=["POST"])
def chart_generate_stream():
    """스트리밍 차트 생성 (SSE) — 실시간으로 차트가 나타남"""
    data = request.json
    transcript = data.get("transcript", "")
    consultant = data.get("consultant", "")
    consultation_type = data.get("consultationType", "auto")
    chart_style = data.get("chartStyle", "detailed")
    customer_profile = data.get("customerProfile")

    if not transcript.strip():
        return jsonify({"error": "상담 내용이 없습니다"})

    print(f"[Chart Stream] 스트리밍 차트 생성 요청 - 텍스트 {len(transcript)}자, 스타일: {chart_style}")

    def generate():
        full_text = ""
        try:
            for chunk in generate_chart_stream(
                transcript=transcript,
                customer_profile=customer_profile,
                consultation_type=consultation_type,
                consultant=consultant,
                chart_style=chart_style,
            ):
                full_text += chunk
                yield f"data: {json.dumps({'text': chunk}, ensure_ascii=False)}\n\n"

            summary = extract_summary(full_text)
            yield f"data: {json.dumps({'done': True, 'summary': summary}, ensure_ascii=False)}\n\n"
            print(f"[Chart Stream] 생성 완료 - 차트 {len(full_text)}자")
        except Exception as e:
            print(f"[Chart Stream] 생성 실패: {e}")
            yield f"data: {json.dumps({'error': str(e)}, ensure_ascii=False)}\n\n"

    return Response(
        stream_with_context(generate()),
        content_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Access-Control-Allow-Origin": "*",
        }
    )


@app.route("/interpret/translate", methods=["POST"])
def interpret_translate():
    data = request.json
    result = translate(
        text=data["text"],
        source_lang=data.get("sourceLang", "ko"),
        target_lang=data["targetLang"],
        speaker_role=data.get("speakerRole", "doctor"),
        custom_prompt=data.get("customPrompt"),
    )
    return jsonify(result)


@app.route("/interpret/chart", methods=["POST"])
def interpret_chart():
    data = request.json
    translated = translate_chart(
        chart_text=data["chart"],
        target_lang=data["targetLang"],
    )
    return jsonify({"translatedChart": translated})


if __name__ == "__main__":
    port = int(os.getenv("AI_PORT", 8081))
    print(f"\n  AI 서버: http://localhost:{port}")
    print()
    app.run(host="0.0.0.0", port=port, debug=False)
